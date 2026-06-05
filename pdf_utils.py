import io
import re
import os
import urllib.request

import altair as alt
import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Register a Unicode CJK-capable font for Chinese glyphs.
def _register_font(path, name):
    try:
        pdfmetrics.registerFont(TTFont(name, path))
        return True
    except Exception:
        return False

# Try to find common CJK fonts on the system
DEFAULT_FONT = None
DEFAULT_BOLD_FONT = None
DEFAULT_ITALIC_FONT = None

system_candidates = [
    '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
    '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.otf',
    '/usr/share/fonts/truetype/NotoSansCJKtc-Regular.otf',
    '/usr/share/fonts/truetype/NotoSansCJKsc-Regular.otf',
    '/usr/share/fonts/truetype/arphic/ukai.ttc',
    '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',
    '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',
    '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
]

found = None
for p in system_candidates:
    if os.path.exists(p):
        low = p.lower()
        # accept only likely CJK-capable fonts (avoid Dejavu which lacks Chinese glyphs)
        if any(k in low for k in ('noto', 'cjk', 'wqy', 'arphic', 'sourcehan', 'ukai', 'unifont')):
            found = p
            break

if found:
    # prefer a clear name for registration
    base_name = 'CustomCJK'
    if _register_font(found, base_name):
        DEFAULT_FONT = base_name
        DEFAULT_BOLD_FONT = base_name
        DEFAULT_ITALIC_FONT = base_name

# Fallback: try to download a Noto Sans CJK subset (from GitHub raw) into workspace
if not DEFAULT_FONT:
    try:
        noto_url = 'https://github.com/googlefonts/noto-cjk/raw/main/Sans/OTF/SimplifiedChinese/NotoSansSC-Regular.otf'
        target = '/tmp/NotoSansSC-Regular.otf'
        if not os.path.exists(target):
            urllib.request.urlretrieve(noto_url, target)
        if _register_font(target, 'NotoSansSC'):
            DEFAULT_FONT = 'NotoSansSC'
            DEFAULT_BOLD_FONT = 'NotoSansSC'
            DEFAULT_ITALIC_FONT = 'NotoSansSC'
    except Exception:
        DEFAULT_FONT = None

# Final fallback to DejaVuSans or Helvetica
if not DEFAULT_FONT:
    try:
        pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
        DEFAULT_FONT = 'DejaVuSans'
        DEFAULT_BOLD_FONT = 'DejaVuSans-Bold'
        DEFAULT_ITALIC_FONT = 'DejaVuSans'
    except Exception:
        DEFAULT_FONT = 'Helvetica'
        DEFAULT_BOLD_FONT = 'Helvetica-Bold'
        DEFAULT_ITALIC_FONT = 'Helvetica'

@st.cache_data
def extract_item_analysis(file_bytes):
    row_pattern = re.compile(
        r'^(.*?)\s+(\d+)\s+(\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+%)\s+(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+%)\s+(\d+\.\d+)\s*([+-]?\d+\.\d+)\s*'
    )
    extracted_data = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            for line in text.split('\n'):
                clean_line = " ".join(line.split())
                match = row_pattern.search(clean_line)
                if match:
                    extracted_data.append(match.groups()[:11])

    columns = [
        "Item", "Max Mark", "Your school Attm. No.", "Your school Attem. %", "Your school Mean",
        "Your school Mean %", "Your school SD", "Day schools Attem. %", "Day schools Mean",
        "Day schools Mean %", "Day schools SD"
    ]
    df = pd.DataFrame(extracted_data, columns=columns)
    df.insert(0, "row_index", range(1, len(df) + 1))
    numeric_cols = [
        "Max Mark", "Your school Attm. No.", "Your school Attem. %", "Your school Mean",
        "Your school SD", "Day schools Attem. %", "Day schools Mean", "Day schools SD"
    ]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    pct_cols = ["Your school Mean %", "Day schools Mean %"]
    for col in pct_cols:
        df[col] = df[col].str.replace('%', '').astype(float) / 100
    return df


@st.cache_data
def extract_mcq_analysis(file_bytes):
    mcq_data = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            lines = text.split('\n')
            current_question = None
            correct_answer = None
            question_answers = {}
            for line in lines:
                q_match = re.match(r'^(\d+\([ivx]+\)|\d+)\s+貴校', line.strip())
                if q_match:
                    if current_question and question_answers:
                        row = {'Question Number': current_question, 'Corr. Ans': correct_answer}
                        for opt in ['A', 'B', 'C', 'D']:
                            row[f'Your school {opt}_No.'] = question_answers.get(f'{opt}_your', '0')
                            row[f'Day schools {opt}_No.'] = question_answers.get(f'{opt}_day', '0')
                        mcq_data.append(row)
                    current_question = q_match.group(1)
                    question_answers = {}
                    correct_answer = None

                answer_match = re.match(r'^([ABCD])\s+()?\s*(\d+)\s+[\d.]+\s+([\d,]+)', line.strip())
                if answer_match and current_question:
                    option = answer_match.group(1)
                    has_marker = answer_match.group(2) is not None
                    your_no = answer_match.group(3)
                    day_no = answer_match.group(4).replace(',', '')
                    if has_marker:
                        correct_answer = option
                    question_answers[f'{option}_your'] = your_no
                    question_answers[f'{option}_day'] = day_no

            if current_question and question_answers:
                row = {'Question Number': current_question, 'Corr. Ans': correct_answer}
                for opt in ['A', 'B', 'C', 'D']:
                    row[f'Your school {opt}_No.'] = question_answers.get(f'{opt}_your', '0')
                    row[f'Day schools {opt}_No.'] = question_answers.get(f'{opt}_day', '0')
                mcq_data.append(row)

    df = pd.DataFrame(mcq_data)
    if not df.empty:
        column_order = [
            'Question Number', 'Corr. Ans',
            'Your school A_No.', 'Your school B_No.', 'Your school C_No.', 'Your school D_No.',
            'Day schools A_No.', 'Day schools B_No.', 'Day schools C_No.', 'Day schools D_No.'
        ]
        df = df[column_order]
        df.insert(0, "row_index", range(1, len(df) + 1))
        for col in df.columns:
            if '_No.' in col:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
    return df


@st.cache_data
def extract_latest_dse_total_data(file_bytes):
    target_grades = ['5**', '5*+', '5+', '4+', '3+', '2+', '1+', 'UNCL', '出席 Sat']
    results = []
    subject_name = "未知科目"
    exam_year = "未知年份"
    attendance_ys = None
    attendance_ds = None
    expected_grade_count = len(target_grades) - 1

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            if "總數" in text and "貴校" in text and "5**" in text:
                lines = text.split('\n')

                for i, line in enumerate(lines):
                    if "HKDSE 20" in line and exam_year == "未知年份":
                        exam_year = line.replace("HKDSE", "").strip()

                for i, line in enumerate(lines):
                    if ("總數 Total" in line or "總數" in line) and subject_name == "未知科目":
                        if i >= 2 and "Category" not in lines[i-2] and "學科" not in lines[i-2] and "results" not in lines[i-2]:
                            subject_name = lines[i-2].strip()
                        elif i >= 1:
                            subject_name = lines[i-1].strip()

                in_total_section = False
                for line in lines:
                    if "總數 Total" in line or "總數" in line:
                        in_total_section = True
                    elif "男生 Male" in line or "女生 Female" in line:
                        in_total_section = False

                    if in_total_section:
                        clean_line = line.replace(',', '')
                        for grade in target_grades:
                            if clean_line.startswith(grade + " "):
                                parts = clean_line.split(grade)
                                if len(parts) >= 3:
                                    ys_numbers = parts[1].strip().split()
                                    ds_numbers = parts[2].strip().split()
                                    if ys_numbers and ds_numbers:
                                        if grade == '出席 Sat':
                                            attendance_ys = int(ys_numbers[-1])
                                            attendance_ds = int(ds_numbers[-1])
                                        elif not any(r['等級'] == grade for r in results):
                                            results.append({
                                                '等級': grade,
                                                '貴校': int(ys_numbers[-1]),
                                                '日校': int(ds_numbers[-1])
                                            })
                                break
                if len(results) == expected_grade_count and attendance_ys is not None and attendance_ds is not None:
                    break

    df = pd.DataFrame(results)
    if not df.empty:
        df['等級'] = pd.Categorical(df['等級'], categories=[g for g in target_grades if g != '出席 Sat'], ordered=True)
        df = df.sort_values('等級').reset_index(drop=True)
    return df, subject_name, exam_year, attendance_ys, attendance_ds


def convert_df_to_excel(df, sheet_name):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return output.getvalue()


def convert_df_to_styled_excel(df, style_map=None, sheet_name="Overview"):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                if isinstance(value, int):
                    cell.number_format = "0"
                else:
                    cell.number_format = "0.00"
            if style_map:
                cell_style = style_map.get((row_idx - 2, col_idx - 1), {})
                if cell_style.get("fill"):
                    color = cell_style["fill"].lstrip("#")
                    if len(color) == 6:
                        color = f"FF{color}"
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                if cell_style.get("font_color") or cell_style.get("bold") or cell_style.get("italic"):
                    font_color = cell_style.get("font_color")
                    if isinstance(font_color, str) and font_color.startswith("#"):
                        font_color = font_color.lstrip("#")
                    cell.font = Font(
                        color=font_color,
                        bold=cell_style.get("bold", False),
                        italic=cell_style.get("italic", False)
                    )

    for col_idx, header in enumerate(headers, start=1):
        max_length = len(str(header))
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 50)

    wb.save(output)
    return output.getvalue()


def convert_df_to_pdf(df, style_map=None, title="Overview Table"):
    output = io.BytesIO()
    left_margin = 20
    right_margin = 20
    top_margin = 20
    bottom_margin = 20
    page_width, _ = landscape(A4)
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=top_margin,
        bottomMargin=bottom_margin,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="Title",
        parent=styles["Heading2"],
        fontName=DEFAULT_FONT,
        fontSize=14,
        leading=18,
        alignment=TA_CENTER,
    )
    story = [Paragraph(title, title_style), Spacer(1, 12)]

    # Determine which columns are numeric (for wrap text decision)
    numeric_cols = set()
    for col_idx, col_name in enumerate(df.columns):
        col_data = df.iloc[:, col_idx]
        is_numeric = pd.api.types.is_numeric_dtype(col_data)
        if is_numeric:
            numeric_cols.add(col_idx)
    
    # Build data with wrapped text for non-numeric columns (with UTF-8 encoding)
    data_style = getSampleStyleSheet()
    wrap_style = ParagraphStyle(
        name="WrapText",
        parent=data_style["Normal"],
        fontName=DEFAULT_FONT,
        fontSize=9,
        leading=11,
        alignment=TA_LEFT,
        wordWrap='CJK',  # Enable CJK word wrapping for Chinese text
    )
    
    header_style = ParagraphStyle(
        name="HeaderText",
        parent=data_style["Normal"],
        fontName=DEFAULT_FONT,
        fontSize=9,
        leading=11,
        alignment=TA_CENTER,
        wordWrap='CJK',
    )
    
    # Header row
    data = []
    header_row = []
    for col_name in df.columns:
        header_row.append(Paragraph(str(col_name), header_style))
    data.append(header_row)
    
    # Data rows
    for row in df.itertuples(index=False):
        data_row = []
        for col_idx, value in enumerate(row):
            value_str = str(value) if value is not None else ""
            if col_idx in numeric_cols:
                data_row.append(value_str)
            else:
                data_row.append(Paragraph(value_str, wrap_style))
        data.append(data_row)

    available_width = page_width - left_margin - right_margin
    estimated_widths = []
    for col_idx in range(len(df.columns)):
        max_len = max(len(str(df.columns[col_idx])), *(len(str(row[col_idx])) for row in df.itertuples(index=False)))
        estimated_widths.append(min(max(40, max_len * 6), 160))

    total_width = sum(estimated_widths)
    if total_width > available_width:
        scale = available_width / total_width
        col_widths = [w * scale for w in estimated_widths]
    else:
        col_widths = estimated_widths

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table_style = [
        ("FONTNAME", (0, 0), (-1, -1), DEFAULT_FONT),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D3D3D3")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]

    if style_map:
        for (row_idx, col_idx), style in style_map.items():
            table_row = row_idx + 1
            table_col = col_idx
            if style.get("fill"):
                table_style.append(("BACKGROUND", (table_col, table_row), (table_col, table_row), colors.HexColor(style["fill"])))
            if style.get("font_color"):
                table_style.append(("TEXTCOLOR", (table_col, table_row), (table_col, table_row), colors.HexColor(style["font_color"])))
            if style.get("bold"):
                table_style.append(("FONTNAME", (table_col, table_row), (table_col, table_row), DEFAULT_BOLD_FONT))
            if style.get("italic"):
                table_style.append(("FONTNAME", (table_col, table_row), (table_col, table_row), DEFAULT_ITALIC_FONT))
    table.setStyle(TableStyle(table_style))
    story.append(table)
    doc.build(story)
    return output.getvalue()
